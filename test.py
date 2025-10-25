import os
import torch
from PIL import Image

import clip
import openpyxl
from util import *
# assume remove_duplicates is defined earlier or include it here:
def remove_duplicates(strings):
    seen = set()
    result = []
    for s in strings:
        if s not in seen:
            seen.add(s)
            result.append(s)
    return result

device = "cuda" if torch.cuda.is_available() else "cpu"

model, preprocess = clip.load("ViT-B/16", device=device)
model.load_state_dict(torch.load(r"models/with_crack_descr_700.pth", map_location=device))
model.eval()

new_path = r"Img_dataset"
imgs = []
true_labels = []            # true labels per image (preserve order)
candidate_labels = []       # pool of text prompts to compare against

xl_file = r"testCLIP.xlsx"
wb = openpyxl.load_workbook(xl_file)
ws = wb.active    

# gather entries from xl file
for i in range(1, ws.max_row + 1):
    path = ws.cell(i, 1).value
    if not path:
        continue
    path = str(path)
    img_name = path.split("/")[-1]
    value = ws.cell(i, 2).value
    if value is None:
        continue
    value = str(value).lower().strip()
    full_img_path = os.path.join(new_path, img_name)
    value = combine_descriptions(ws, i)
    if (value is None):
        continue
    if os.path.exists(full_img_path):
        imgs.append(full_img_path)
        value = "an image of " + value
        true_labels.append(value)
        candidate_labels.append(value)  # add the true label to candidate pool

# some garbage prompts to add to the candidate pool (distractors)
# candidate_labels.append("Dr.Limestone cracked their back on the brick wall")
# candidate_labels.append("Stepped into the crack under the balcony")

# deduplicate candidate pool (preserves order)
candidate_labels = remove_duplicates(candidate_labels)

# sanity check
if len(imgs) == 0:
    raise SystemExit("No images found. Check new_path and Excel file paths.")

# Pre-tokenize and encode all candidate labels ONCE (speeds up evaluation)
text_inputs = torch.cat([clip.tokenize(c) for c in candidate_labels]).to(device)

with torch.no_grad():
    text_features = model.encode_text(text_inputs)
    text_features /= text_features.norm(dim=-1, keepdim=True)

num_correct = 0

for idx, img_path in enumerate(imgs):
    img = Image.open(img_path).convert("RGB")

    image_input = preprocess(img).unsqueeze(0).to(device)

    with torch.no_grad():
        image_features = model.encode_image(image_input)
    image_features /= image_features.norm(dim=-1, keepdim=True)

    # similarity over the deduplicated candidate pool
    logits_per_image = (100.0 * image_features @ text_features.T)
    probs = logits_per_image.softmax(dim=-1)

    values, indices = probs[0].topk(min(3, len(candidate_labels)))  # top-3 predictions
    if (candidate_labels[indices[0].item()] == true_labels[idx]):
        print(f"\n CORRECT Image: {img_path}")
    else:
        print(f"\n NOT CORRECT Image: {img_path}")

    print("Top predictions:")
    for rank, (val, idx_label) in enumerate(zip(values, indices), start=1):
        print(f"  {rank:>2}. {candidate_labels[idx_label]:>60s} : {100 * val.item():.2f}%")

    # check top-1 against the true label for this image
    top1_index = indices[0].item()
    predicted_label = candidate_labels[top1_index]
    actual_label = true_labels[idx]

    if predicted_label == actual_label:
        num_correct += 1

    print(f"Actual description: {actual_label}")

# final accuracy
total = len(imgs)
accuracy = num_correct / total if total > 0 else 0.0
print(f"\ncorrect: {num_correct}  total: {total}  percent: {accuracy:.4f}")
