# All CLIP is trying to do is map a text input and an image into the same latent space so that they are related. 
# This means if that if trained well, the image and description should be mapped to the same latent space. 
# If trained well, it should be able to take in an iamge and output a description of that image.
# heres the article where i got the majority of the code from: https://medium.com/aimonks/a-guide-to-fine-tuning-clip-models-with-custom-data-6c7c0d1416fb

import torch
import torch.nn as nn
from torch.utils.data import DataLoader

import clip
import tqdm
import openpyxl
import os
from image_tileset import image_title_dataset
from util import *

DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
NUM_EPOCHS = 701
START_EPOCH = 0
if (START_EPOCH >=  NUM_EPOCHS):
    print("reconfig bounds of epochs")
    exit(1)
model, preprocess = clip.load("ViT-B/16", device=DEVICE)
# model.load_state_dict(torch.load(r"models/clip_model_finetuned_400.pth", map_location=DEVICE))


def convert_models_to_fp32(model): # used to convert weights to float 32 for more precise updating
    for p in model.parameters():
        p.data = p.data.float()
        p.grad.data = p.grad.data.float()


# xl loading stuff
xl_file = r"Training.xlsx"
wb = openpyxl.load_workbook(xl_file)
sheets = wb.sheetnames
ws = wb[sheets[0]]



# gathering entries from xl file
imgs = []
text = []
training_file = r"Img_dataset" # Training images folder is in the drive

not_found = 0
# loading up the image names and their descriptions
end_line = 473 #  ws.max_row + 1 # 
for i in range(1, end_line): 
    path = str(ws.cell(i, 1).value)
    img_name = path.split("/")[-1]
    value = str(ws.cell(i, 2).value)
    if (os.path.exists(training_file + "/" + img_name)):
        extra_descriptions = combine_descriptions(ws, i) # here is where im combing the extra descriptions i annotated. Removing may lead to better
        if (extra_descriptions is not None):
            value = extra_descriptions
            text.append("an image of "+value)
            imgs.append(training_file + "/" + img_name)

        
        
print(f"amount of images: {len(imgs)}")
print(f"amount of images: {len(text)}")
print(f"not found: {not_found}")


optimizer = torch.optim.Adam(model.parameters(), lr=5e-5,betas=(0.9,0.98),eps=1e-6,weight_decay=0.2)
loss_img = nn.CrossEntropyLoss()
loss_txt = nn.CrossEntropyLoss()

dataset = image_title_dataset(imgs, text, preprocess) # Here we create an object that will take care of loading an image
# the obejct is mainly so we can pass something into the DataLoader class
train_dataloader = DataLoader(dataset, batch_size=20, shuffle=True) # try reducing batch size if you run out of memory

for epoch in range(START_EPOCH, NUM_EPOCHS):
    pbar = tqdm.tqdm(train_dataloader, total=len(train_dataloader))
    for batch in pbar:
        optimizer.zero_grad()

        images,texts = batch

        images= images.to(DEVICE)
        texts = texts.to(DEVICE)

        # Forward pass
        logits_per_image, logits_per_text = model(images, texts)
        # Compute loss
        ground_truth = torch.arange(len(images),dtype=torch.long,device=DEVICE)
        total_loss = (loss_img(logits_per_image,ground_truth) + loss_txt(logits_per_text,ground_truth))/2

        # Backward pass
        total_loss.backward()
        if DEVICE == "cpu":
            optimizer.step()
        else :
            convert_models_to_fp32(model) # converting weight numbers to float 32 for precise updating 
            optimizer.step()
            clip.model.convert_weights(model)

        pbar.set_description(f"Epoch {epoch}/{NUM_EPOCHS}, Loss: {total_loss.item():.4f}")
        if (epoch % 50 ==0 and epoch != 0):
            torch.save(model.state_dict(), f"models/with_crack_descr_{epoch}.pth")
